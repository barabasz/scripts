import pyexcel
from openpyxl import load_workbook

filename = "test.xls"
records = pyexcel.iget_records(file_name=filename)
names = [record['name'] for record in records]
print(type(names), names)

filename = "test.xlsx"
book = load_workbook(filename=filename)
first_sheet = book.sheetnames[0]
sheet = book[first_sheet]
column = sheet['B']
names = [cell.value for cell in column if cell.row != 1]
print(type(names), names)
