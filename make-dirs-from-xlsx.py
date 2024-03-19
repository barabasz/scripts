import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

sheet_index = 0
folders_col_name = "Folder"
data_file = ".\\filename.xlsx"
root_path = os.path.dirname(os.path.realpath(__file__))

workbook = load_workbook(filename=data_file, data_only=True)
sheetname = workbook.sheetnames[sheet_index]
sheet = workbook[sheetname]
headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
colum_number = [col[0].value for col in sheet.iter_cols()].index(folders_col_name) + 1
column_letter = get_column_letter(colum_number)
column = sheet[column_letter]
folders = [cell.value for cell in column if cell.row != 1]

counter = 0

for folder in folders:
    os.mkdir(os.path.join(root_path, folder))
    counter += 1

print(f"{counter} folder(s) created")
